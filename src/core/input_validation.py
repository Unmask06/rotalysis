"""This module contains the Pydantic model for the input configuration file"""

import logging
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from pydantic import BaseModel, field_validator, model_validator

logger = logging.getLogger("Input Validation")


class PipSimInput(BaseModel):
    """Pydantic model for the input configuration file."""

    FOLDER_DIRECTORY: Path
    MODEL_FILENAME: str
    EXCEL_FILE: str
    PIPSIM_INPUT_SHEET: str
    CONDITIONS_SHEET: str
    SOURCE_NAME: str
    PUMP_NAME: List[str]
    STRAINER_NAME: List[str]

    @field_validator("FOLDER_DIRECTORY")
    @classmethod
    def check_folder_directory(cls, v):
        if not Path(v).is_dir():
            raise ValueError("Folder directory does not exist")
        return v

    @model_validator(mode="after")
    def check_model_filename(self):
        folder_directory = self.FOLDER_DIRECTORY
        model_filename = self.MODEL_FILENAME
        if (
            not Path(folder_directory / model_filename).is_file()
            or not Path(model_filename).is_file()
        ):
            raise ValueError("Model file does not exist")
        return self

    @model_validator(mode="after")
    def check_excel_file(self):
        folder_directory = self.FOLDER_DIRECTORY
        excel_file = self.EXCEL_FILE
        if (
            not Path(folder_directory / excel_file).is_file()
            or not Path(excel_file).is_file()
        ):
            raise ValueError("Excel file does not exist")
        return self

    @model_validator(mode="after")
    def check_pipsim_input_sheet(self):
        folder_directory = self.FOLDER_DIRECTORY
        excel_file = self.EXCEL_FILE
        excel_file_path = folder_directory / excel_file
        wb = load_workbook(excel_file_path, read_only=True)
        if self.PIPSIM_INPUT_SHEET not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{self.PIPSIM_INPUT_SHEET}' does not exist in {excel_file}"
            )
        return self
